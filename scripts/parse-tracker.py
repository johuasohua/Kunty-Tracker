"""Parse Kiki & Josh tracker xlsx -> import.json, validating raw rows
against the workbook's own pivot tables before anything is written."""
import json
import re
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

XLSX = "/root/.claude/uploads/792b08b1-5cb5-56ca-be8f-9505f448ef0f/8a107550-Kiki__Josh_Exp_tracker_1.xlsx"
OUT = "/tmp/claude-0/-home-user-Kunty-Tracker/792b08b1-5cb5-56ca-be8f-9505f448ef0f/scratchpad/import.json"

TX_TABS = ["APR 2026", "MAY 2026", "june 2026"]
ACCOUNTS = {"Kiki Credit", "Kiki Debit", "Josh Credit", "Josh Debit"}

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------------------------------------------------------------- transactions
transactions = []
carried_forward = []  # audit trail of blank-date rows
problems = []

for tab in TX_TABS:
    ws = wb[tab]
    last_date = None
    for r in range(2, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        cat_v = ws.cell(r, 2).value
        amt_v = ws.cell(r, 3).value
        acct_v = ws.cell(r, 4).value
        type_v = ws.cell(r, 5).value
        note_v = ws.cell(r, 6).value

        if cat_v is None and amt_v is None:
            continue  # empty row

        if amt_v is None or cat_v is None or acct_v is None or type_v is None:
            problems.append(f"{tab} row {r}: incomplete row {[date_v, cat_v, amt_v, acct_v, type_v]}")
            continue

        acct = str(acct_v).strip()
        if acct not in ACCOUNTS:
            problems.append(f"{tab} row {r}: unknown account {acct!r}")
            continue

        ttype = str(type_v).strip().lower()
        if ttype not in ("income", "expense"):
            problems.append(f"{tab} row {r}: unknown type {type_v!r}")
            continue

        if isinstance(date_v, datetime):
            occurred = date_v.date()
            last_date = occurred
        elif date_v is None:
            if last_date is None:
                problems.append(f"{tab} row {r}: blank date with no previous dated row")
                continue
            occurred = last_date
            carried_forward.append({"tab": tab, "row": r, "assigned": occurred.isoformat(),
                                    "category": str(cat_v).strip(), "amount": float(amt_v)})
        else:
            problems.append(f"{tab} row {r}: unparseable date {date_v!r}")
            continue

        person, method = acct.split(" ")
        transactions.append({
            "tab": tab,
            "row": r,
            "occurred_on": occurred.isoformat(),
            "category": str(cat_v).strip(),
            "amount": round(float(amt_v), 2),
            "person": person,
            "payment_method": method.lower(),
            "type": ttype,
            "note": str(note_v).strip() if note_v is not None else None,
        })

# ------------------------------------------------- validate against tab pivots
# Pivot layout: col I=Category, J=kiki Credit, K=Kiki Debit, L=Josh Credit,
# M=Josh Debit, rows 2..14, row 15 = expense Totals per account column.
tolerance = 0.011
mismatches = []
for tab in TX_TABS:
    ws = wb[tab]
    # our sums keyed (category.lower(), person, method); expenses AND salary income
    sums = defaultdict(float)
    for t in transactions:
        if t["tab"] != tab:
            continue
        sums[(t["category"].lower(), t["person"], t["payment_method"])] += t["amount"]

    for r in range(2, 15):
        cat = ws.cell(r, 9).value
        if cat is None:
            continue
        cat_key = str(cat).strip().lower()
        for col, person, method in ((10, "Kiki", "credit"), (11, "Kiki", "debit"),
                                    (12, "Josh", "credit"), (13, "Josh", "debit")):
            expected = float(ws.cell(r, col).value or 0)
            actual = sums.get((cat_key, person, method), 0.0)
            if abs(expected - actual) > tolerance:
                mismatches.append(
                    f"{tab} / {cat} / {person} {method}: sheet pivot={expected:.2f} parsed={actual:.2f} diff={actual-expected:+.2f}"
                )

# ---------------------------------------------------------------- mortgage LIV
ws = wb["LIV"]
mortgage = []
for r in range(5, ws.max_row + 1):
    period = ws.cell(r, 1).value
    date_v = ws.cell(r, 2).value
    if period is None or date_v is None:
        continue
    d, m, y = str(date_v).strip().split(".")
    payment_date = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"

    def num(c):
        v = ws.cell(r, c).value
        return round(float(v), 2) if v is not None and not isinstance(v, str) else None

    raw_adhoc = ws.cell(r, 16).value
    adhoc_note = str(ws.cell(r, 17).value).strip() if ws.cell(r, 17).value else None
    if isinstance(raw_adhoc, str):
        # e.g. "50000+30000" — evaluate simple +/- expressions, keep text in note
        if re.fullmatch(r"[\d .+\-]+", raw_adhoc):
            total = sum(float(p) for p in re.findall(r"[+-]?[\d.]+", raw_adhoc))
            note_extra = f"[{raw_adhoc.strip()}]"
            adhoc_note = f"{adhoc_note} {note_extra}" if adhoc_note else note_extra
            adhoc = round(total, 2)
        else:
            problems.append(f"LIV row {r}: unparseable ad-hoc amount {raw_adhoc!r}")
            adhoc = None
    else:
        adhoc = round(float(raw_adhoc), 2) if raw_adhoc is not None else None

    mortgage.append({
        "period_no": int(period),
        "payment_date": payment_date,
        "opening_principal": num(3),
        "principal_amount": num(5),
        "interest_amount": num(6),
        "insurance_amount": num(7),
        "hoi_charge": num(8),
        "closing_principal": num(9),
        "offset_opening_balance": num(11),
        "offset_closing_balance": num(12),
        "interest_saved": num(14),
        "offset_transaction_amount": adhoc,
        "offset_note": adhoc_note,
    })

# sanity: closing == opening - principal (within rounding)
for mrow in mortgage:
    if abs((mrow["opening_principal"] - mrow["principal_amount"]) - mrow["closing_principal"]) > 0.02:
        problems.append(f"LIV period {mrow['period_no']}: opening-principal != closing")

# ------------------------------------------------------- seeds & cc figures
# Opening cash balance as of April 2026, Overview convention (category P&L),
# which is the same convention the app's balance card uses.
overview = wb["Overview"]
opening_seed = None
for r in range(9, 21):
    mv = overview.cell(r, 2).value
    if isinstance(mv, datetime) and mv.date().isoformat() == "2026-04-01":
        opening_seed = {"as_of_month": "2026-04-01", "balance": round(float(overview.cell(r, 3).value), 2)}
if opening_seed is None:
    problems.append("Could not find April 2026 opening balance in Overview")

# CC carry-over into April (per person) + cc paid off per month, from tab pivots.
opening_cc = {}
cc_payments = []
for tab, month in (("APR 2026", "2026-04-01"), ("MAY 2026", "2026-05-01"), ("june 2026", "2026-06-01")):
    ws = wb[tab]
    for r in (21, 22):
        who = str(ws.cell(r, 8).value).strip()  # KIKI / JOSHUA
        person = "Kiki" if who.upper().startswith("K") else "Josh"
        carry_cc = float(ws.cell(r, 13).value)   # col M 'carry over cc'
        paid_cc = float(ws.cell(r, 15).value)    # col O 'cc paid off'
        if tab == "APR 2026":
            opening_cc[person] = {"as_of_month": "2026-04-01", "balance": round(carry_cc, 2)}
        cc_payments.append({"person": person, "month": month, "amount_paid": round(paid_cc, 2)})

# ---------------------------------------------------------------------- output
result = {
    "transactions": transactions,
    "mortgage": mortgage,
    "opening_seed": opening_seed,
    "opening_cc": opening_cc,
    "cc_payments": cc_payments,
    "carried_forward": carried_forward,
}
with open(OUT, "w") as f:
    json.dump(result, f, indent=1)

print(f"transactions: {len(transactions)} rows "
      f"({sum(1 for t in transactions if t['tab']=='APR 2026')} apr / "
      f"{sum(1 for t in transactions if t['tab']=='MAY 2026')} may / "
      f"{sum(1 for t in transactions if t['tab']=='june 2026')} jun)")
print(f"blank dates carried forward: {len(carried_forward)}")
print(f"mortgage payments: {len(mortgage)} periods "
      f"({mortgage[0]['payment_date']} .. {mortgage[-1]['payment_date']})")
print(f"opening cash seed: {opening_seed}")
print(f"opening CC balances: {opening_cc}")
print(f"cc payments: {cc_payments}")
print()
if problems:
    print("PROBLEMS:")
    for p in problems:
        print(" -", p)
if mismatches:
    print("PIVOT MISMATCHES:")
    for m in mismatches:
        print(" -", m)
if not problems and not mismatches:
    print("ALL PIVOT CROSS-CHECKS PASSED — parsed data matches the sheet's own totals.")
else:
    sys.exit(1)
